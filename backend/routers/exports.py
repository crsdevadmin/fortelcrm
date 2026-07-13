# backend/routers/exports.py
"""
Excel export endpoints.
  GET /exports/sales?year=&month=&associate_id=   → monthly sales by doctor
  GET /exports/rep-activity?year=&month=          → visits + sales per rep per week
  GET /exports/doctor-master                      → full doctor list with ROI data
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import datetime, timedelta
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from sqlalchemy import text

from ..database import get_db
from ..models.models import SalesEntry, Doctor, Product, User, VisitLog

router = APIRouter(prefix="/exports", tags=["Exports"])


# ── Helpers ──────────────────────────────────────────────────────────────────

def _header_style(ws, row, cols, bg="1A3A1A", fg="FFFFFF"):
    """Apply dark header style to a row."""
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, size=10)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, padding=4):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + padding


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _subordinate_ids(db, user_id):
    result = {user_id}
    queue  = [user_id]
    while queue:
        cur = queue.pop()
        subs = db.query(User.id).filter(User.reports_to_id == cur).all()
        for (sid,) in subs:
            if sid not in result:
                result.add(sid)
                queue.append(sid)
    return result


# ── 1. Monthly Sales Export ───────────────────────────────────────────────────

@router.get("/sales")
def export_monthly_sales(
    year:         int            = Query(...),
    month:        int            = Query(...),
    associate_id: Optional[int]  = Query(None),
    viewer_id:    Optional[int]  = Query(None),
    db:           Session        = Depends(get_db),
):
    """Download monthly sales entries as Excel."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    q = db.query(SalesEntry).filter(
        SalesEntry.year  == year,
        SalesEntry.month == month,
    )
    if associate_id:
        q = q.filter(SalesEntry.associate_id == associate_id)
    elif viewer_id:
        visible = _subordinate_ids(db, viewer_id)
        q = q.filter(SalesEntry.associate_id.in_(visible))

    rows = q.order_by(SalesEntry.sale_date, SalesEntry.doctor_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sales {year}-{month:02d}"

    headers = ["Date", "Rep", "Doctor", "Hospital", "City", "Specialty",
               "Product", "Qty", "Value (₹)", "Approved"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    for r_idx, e in enumerate(rows, 2):
        doc  = e.doctor  if e.doctor  else None
        prod = e.product if e.product else None
        rep  = db.query(User).filter(User.id == e.associate_id).first()
        ws.append([
            e.sale_date or f"{year}-{month:02d}-{(e.week or 1):02d}",
            rep.name  if rep  else e.associate_id,
            doc.name  if doc  else e.doctor_id,
            doc.hospital if doc else "",
            doc.city     if doc else "",
            doc.specialty if doc else "",
            prod.name if prod else e.product_id,
            e.qty   or 0,
            round(e.value or 0, 2),
            "Yes" if e.is_approved else "No",
        ])

    # Value column as number
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    _auto_width(ws)

    from calendar import month_abbr
    mn = month_abbr[month]
    return _xlsx_response(wb, f"Fortel_Sales_{mn}{year}.xlsx")


# ── 2. Rep Activity Export ────────────────────────────────────────────────────

@router.get("/rep-activity")
def export_rep_activity(
    year:      int           = Query(...),
    month:     int           = Query(...),
    viewer_id: Optional[int] = Query(None),
    db:        Session       = Depends(get_db),
):
    """Per-rep weekly activity: visits, hospitals, doctors, sales."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}

    visible_ids = _subordinate_ids(db, viewer_id) if viewer_id else None

    # Get all reps in scope
    q = db.query(User).filter(User.role.in_(["rep", "custom", "associate"]))
    if visible_ids:
        q = q.filter(User.id.in_(visible_ids))
    reps = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rep Activity"

    headers = ["Rep", "Role", "Week", "Visit Days", "Unique Hospitals",
               "Unique Doctors", "Sales Entries", "Sales Value (₹)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    # Determine week boundaries for the month
    from calendar import monthrange
    _, days_in_month = monthrange(year, month)
    week_ranges = [
        (1,  7,  "Week 1"),
        (8,  14, "Week 2"),
        (15, 21, "Week 3"),
        (22, days_in_month, "Week 4"),
    ]

    for rep in reps:
        for wk_start, wk_end, wk_label in week_ranges:
            # Build date strings for filtering
            date_from = f"{year}-{month:02d}-{wk_start:02d}"
            date_to   = f"{year}-{month:02d}-{wk_end:02d}"

            # Visits in this week
            visits = db.query(VisitLog).filter(
                VisitLog.associate_id == rep.id,
                func.date(VisitLog.visit_time) >= date_from,
                func.date(VisitLog.visit_time) <= date_to,
            ).all()

            visit_days   = len(set(str(v.visit_time.date()) for v in visits))
            uniq_doctors = len(set(v.doctor_id for v in visits if v.doctor_id))
            uniq_hospitals = len(set(
                v.doctor.hospital for v in visits
                if v.doctor and v.doctor.hospital
            ))

            # Sales entries in this week
            sales_rows = db.query(SalesEntry).filter(
                SalesEntry.associate_id == rep.id,
                SalesEntry.year  == year,
                SalesEntry.month == month,
                SalesEntry.week  >= wk_start,
                SalesEntry.week  <= wk_end,
            ).all()

            sales_entries = len(sales_rows)
            sales_value   = round(sum(s.value or 0 for s in sales_rows), 2)

            if visit_days == 0 and sales_entries == 0:
                continue  # skip empty weeks

            ws.append([
                rep.name, rep.role, wk_label,
                visit_days, uniq_hospitals, uniq_doctors,
                sales_entries, sales_value,
            ])

    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    _auto_width(ws)

    from calendar import month_abbr
    mn = month_abbr[month]
    return _xlsx_response(wb, f"Fortel_RepActivity_{mn}{year}.xlsx")


# ── 3. Doctor Master Export ───────────────────────────────────────────────────

@router.get("/doctor-master")
def export_doctor_master(
    viewer_id: Optional[int] = Query(None),
    db:        Session       = Depends(get_db),
):
    """Full doctor list with ROI data."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}

    q = db.query(Doctor).filter(Doctor.is_active != False)
    doctors = q.order_by(Doctor.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doctor Master"

    headers = [
        "Name", "Hospital", "City", "Specialty", "Category",
        "Expected Multiple", "ROI Grade", "Commercial Model",
        "Manager", "Phone", "State"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    for d in doctors:
        ws.append([
            d.name,
            d.hospital or "",
            d.city     or "",
            d.specialty or d.customer_type or "",
            d.category  or "",
            d.expected_multiple or "",
            d.roi_grade or "",
            d.commercial_model or "",
            d.manager.name if d.manager else "",
            d.phone or "",
            d.state_code or "",
        ])

    ws.freeze_panes = "A2"
    _auto_width(ws)

    return _xlsx_response(wb, "Fortel_DoctorMaster.xlsx")


# ── 4. Weekly Sales JSON (for Regional Sales screen) ─────────────────────────

@router.get("/weekly-sales")
def get_weekly_sales(
    doctor_id: int           = Query(...),
    year:      int           = Query(...),
    month:     int           = Query(...),
    db:        Session       = Depends(get_db),
):
    """Returns product-week grid for one doctor+month. Days 1-7=W1, 8-14=W2, 15-21=W3, 22+=W4."""
    entries = db.query(SalesEntry).filter(
        SalesEntry.doctor_id == doctor_id,
        SalesEntry.year      == year,
        SalesEntry.month     == month,
    ).all()

    def day_to_week(day):
        d = day or 1
        if d <= 7:  return 1
        if d <= 14: return 2
        if d <= 21: return 3
        return 4

    result = {}
    for e in entries:
        pid = e.product_id
        if pid not in result:
            prod = e.product
            result[pid] = {
                "product_id":   pid,
                "product_name": prod.name if prod else f"Product {pid}",
                "gst":          prod.gst  if prod else "5%",
                "rate":         (prod.price or prod.rate or 0) if prod else 0,
                "mrp":          prod.mrp  if prod else 0,
                "w1": 0.0, "w2": 0.0, "w3": 0.0, "w4": 0.0,
            }
        wk = day_to_week(e.week)
        result[pid][f"w{wk}"] += round(e.value or 0, 2)

    products = list(result.values())
    for p in products:
        p["total"] = round(p["w1"] + p["w2"] + p["w3"] + p["w4"], 2)

    grand_total = round(sum(p["total"] for p in products), 2)
    return {"doctor_id": doctor_id, "year": year, "month": month,
            "products": products, "grand_total": grand_total}


# ── 5. Regional Monthly Summary (all doctors for a viewer) ───────────────────

@router.get("/regional-summary")
def get_regional_summary(
    year:      int           = Query(...),
    month:     int           = Query(...),
    viewer_id: Optional[int] = Query(None),
    db:        Session       = Depends(get_db),
):
    """Summary of all doctors that have sales entries for the month, for a given viewer."""
    from ..models.models import Doctor, RepDoctorMapping

    q = db.query(SalesEntry).filter(
        SalesEntry.year  == year,
        SalesEntry.month == month,
    )
    if viewer_id:
        visible = _subordinate_ids(db, viewer_id)
        q = q.filter(SalesEntry.associate_id.in_(visible))

    rows = q.all()

    def day_to_week(day):
        d = day or 1
        if d <= 7:  return 1
        if d <= 14: return 2
        if d <= 21: return 3
        return 4

    # Group by doctor
    docs_map = {}
    for e in rows:
        did = e.doctor_id
        if did not in docs_map:
            doc = e.doctor
            docs_map[did] = {
                "doctor_id":   did,
                "doctor_name": doc.name     if doc else f"Doctor {did}",
                "hospital":    doc.hospital if doc else "",
                "city":        doc.city     if doc else "",
                "w1": 0.0, "w2": 0.0, "w3": 0.0, "w4": 0.0,
            }
        wk = day_to_week(e.week)
        docs_map[did][f"w{wk}"] += round(e.value or 0, 2)

    result = []
    for d in docs_map.values():
        d["total"] = round(d["w1"] + d["w2"] + d["w3"] + d["w4"], 2)
        result.append(d)

    result.sort(key=lambda x: x["total"], reverse=True)
    grand = round(sum(d["total"] for d in result), 2)
    return {"year": year, "month": month, "doctors": result, "grand_total": grand}


# ── 6. Rep Activity JSON (for dashboard screen) ─────────────────────────────── (for dashboard screen) ───────────────────────────────

@router.get("/rep-activity-data")
def get_rep_activity_data(
    year:      int           = Query(...),
    month:     int           = Query(...),
    viewer_id: Optional[int] = Query(None),
    db:        Session       = Depends(get_db),
):
    """JSON version of rep activity — used by the RepActivity dashboard screen."""
    from calendar import monthrange
    visible_ids = _subordinate_ids(db, viewer_id) if viewer_id else None

    q = db.query(User).filter(User.role.in_(["rep", "custom", "associate"]))
    if visible_ids:
        q = q.filter(User.id.in_(visible_ids))
    reps = q.order_by(User.name).all()

    _, days_in_month = monthrange(year, month)
    week_ranges = [
        (1,  7,  "Week 1"),
        (8,  14, "Week 2"),
        (15, 21, "Week 3"),
        (22, days_in_month, "Week 4"),
    ]

    result = []
    for rep in reps:
        rep_weeks = []
        total_visits = 0
        total_sales_val = 0.0

        for wk_start, wk_end, wk_label in week_ranges:
            date_from = f"{year}-{month:02d}-{wk_start:02d}"
            date_to   = f"{year}-{month:02d}-{wk_end:02d}"

            visits = db.query(VisitLog).filter(
                VisitLog.associate_id == rep.id,
                func.date(VisitLog.visit_time) >= date_from,
                func.date(VisitLog.visit_time) <= date_to,
            ).all()

            visit_days     = len(set(str(v.visit_time.date()) for v in visits))
            uniq_doctors   = len(set(v.doctor_id for v in visits if v.doctor_id))
            uniq_hospitals = len(set(
                v.doctor.hospital for v in visits
                if v.doctor and v.doctor.hospital
            ))

            sales_rows = db.query(SalesEntry).filter(
                SalesEntry.associate_id == rep.id,
                SalesEntry.year  == year,
                SalesEntry.month == month,
                SalesEntry.week  >= wk_start,
                SalesEntry.week  <= wk_end,
            ).all()

            sales_entries = len(sales_rows)
            sales_value   = round(sum(s.value or 0 for s in sales_rows), 2)

            total_visits    += visit_days
            total_sales_val += sales_value

            rep_weeks.append({
                "week":            wk_label,
                "visit_days":      visit_days,
                "unique_hospitals": uniq_hospitals,
                "unique_doctors":  uniq_doctors,
                "sales_entries":   sales_entries,
                "sales_value":     sales_value,
            })

        result.append({
            "rep_id":        rep.id,
            "rep_name":      rep.name,
            "role":          rep.role,
            "total_visits":  total_visits,
            "total_sales":   round(total_sales_val, 2),
            "weeks":         rep_weeks,
        })

    return result


# ══════════════════════════════════════════════════════════════════════════════
# REGIONAL / WEEKLY SALES  (no doctor dimension — rep-level product entry)
# ══════════════════════════════════════════════════════════════════════════════

_WEEKLY_TABLE_CREATED = False

def _ensure_weekly_table(db: Session):
    global _WEEKLY_TABLE_CREATED
    if _WEEKLY_TABLE_CREATED:
        return
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS rep_weekly_sales (
            id           SERIAL PRIMARY KEY,
            associate_id INTEGER NOT NULL,
            product_id   INTEGER NOT NULL,
            year         INTEGER NOT NULL,
            month        INTEGER NOT NULL,
            week         INTEGER NOT NULL,
            qty          FLOAT   DEFAULT 0,
            price        FLOAT   DEFAULT 0,
            value        FLOAT   DEFAULT 0,
            submitted_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (associate_id, product_id, year, month, week)
        )
    """))
    db.commit()
    _WEEKLY_TABLE_CREATED = True


class WeeklyItem(BaseModel):
    product_id: int
    w1_qty:   float = 0; w1_price: float = 0
    w2_qty:   float = 0; w2_price: float = 0
    w3_qty:   float = 0; w3_price: float = 0
    w4_qty:   float = 0; w4_price: float = 0
    gst_rate: float = 0.05


class WeeklyEntryPayload(BaseModel):
    associate_id: int
    year:         int
    month:        int
    items:        list[WeeklyItem]


@router.post("/weekly-entry")
def save_weekly_entry(payload: WeeklyEntryPayload, db: Session = Depends(get_db)):
    """Upsert rep weekly product sales (no doctor required)."""
    _ensure_weekly_table(db)

    for item in payload.items:
        for wk in range(1, 5):
            qty   = getattr(item, f"w{wk}_qty",   0) or 0
            price = getattr(item, f"w{wk}_price", 0) or 0
            value = round(qty * price * (1 + item.gst_rate), 2)
            db.execute(text("""
                INSERT INTO rep_weekly_sales
                    (associate_id, product_id, year, month, week, qty, price, value, submitted_at)
                VALUES
                    (:aid, :pid, :yr, :mo, :wk, :qty, :price, :val, NOW())
                ON CONFLICT (associate_id, product_id, year, month, week)
                DO UPDATE SET qty=:qty, price=:price, value=:val, submitted_at=NOW()
            """), {"aid": payload.associate_id, "pid": item.product_id,
                   "yr": payload.year, "mo": payload.month, "wk": wk,
                   "qty": qty, "price": price, "val": value})
    db.commit()
    return {"status": "saved", "items": len(payload.items)}


@router.get("/my-weekly")
def get_my_weekly(
    associate_id: int          = Query(...),
    year:         int          = Query(...),
    month:        int          = Query(...),
    db:           Session      = Depends(get_db),
):
    """Load existing weekly sales for a rep+month."""
    _ensure_weekly_table(db)

    rows = db.execute(text("""
        SELECT r.product_id, p.name AS product_name, p.price AS rate, p.gst, p.mrp,
               r.week, r.qty, r.price, r.value
        FROM   rep_weekly_sales r
        JOIN   products p ON p.id = r.product_id
        WHERE  r.associate_id = :aid AND r.year = :yr AND r.month = :mo
        ORDER  BY p.name, r.week
    """), {"aid": associate_id, "yr": year, "mo": month}).fetchall()

    products_map = {}
    for row in rows:
        pid = row.product_id
        if pid not in products_map:
            products_map[pid] = {
                "product_id":   pid,
                "product_name": row.product_name,
                "rate":  row.rate  or 0,
                "gst":   row.gst   or "5%",
                "mrp":   row.mrp   or 0,
                "w1_qty": 0, "w1_price": 0,
                "w2_qty": 0, "w2_price": 0,
                "w3_qty": 0, "w3_price": 0,
                "w4_qty": 0, "w4_price": 0,
            }
        wk = row.week
        if 1 <= wk <= 4:
            products_map[pid][f"w{wk}_qty"]   = row.qty   or 0
            products_map[pid][f"w{wk}_price"] = row.price or 0

    result = []
    for p in products_map.values():
        total = 0
        for wk in range(1, 5):
            total += (p[f"w{wk}_qty"] or 0) * (p[f"w{wk}_price"] or 0)
        p["total"] = round(total, 2)
        result.append(p)

    return {"associate_id": associate_id, "year": year, "month": month,
            "products": result, "grand_total": round(sum(p["total"] for p in result), 2)}


@router.get("/weekly-history")
def get_weekly_history(
    associate_id: int     = Query(...),
    db:           Session = Depends(get_db),
):
    """List months that have weekly sales data for this rep."""
    _ensure_weekly_table(db)
    rows = db.execute(text("""
        SELECT year, month, SUM(value) AS total
        FROM   rep_weekly_sales
        WHERE  associate_id = :aid
        GROUP  BY year, month
        ORDER  BY year DESC, month DESC
    """), {"aid": associate_id}).fetchall()
    return [{"year": r.year, "month": r.month, "total": round(r.total or 0, 2)} for r in rows]
