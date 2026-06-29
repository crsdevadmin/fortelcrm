"""
Import Telangana CRM Investment Data → investments table
─────────────────────────────────────────────────────────
Sheet mapping:
  First_Image  = April 2026
  Second_Image = May   2026
  Third_Image  = June  2026

Columns: S.No | Doctor | Hospital | Sales | CRM
  CRM = investment amount → investments table

Run: python -m backend.import_telangana_investments
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from difflib import SequenceMatcher

from backend.database import SessionLocal
from backend.models.models import Doctor, User, Investment, InvestmentCategory, InvestmentSubCategory

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "Doctor_Master_Data_All_3_Sheets.xlsx")

SHEET_MONTHS = {
    "First_Image":  (2026, 4),
    "Second_Image": (2026, 5),
    "Third_Image":  (2026, 6),
}

def log(msg): print(msg)

def normalise(s):
    """Lowercase, remove Dr./dr., strip punctuation."""
    s = (s or "").lower()
    for r in ["dr.", "dr ", "dr.", "dr "]:
        s = s.replace(r, "")
    return s.strip()

def name_score(a, b):
    a, b = normalise(a), normalise(b)
    if not a or not b:
        return 0.0
    # exact
    if a == b:
        return 1.0
    # one contained in other
    if a in b or b in a:
        return 0.9
    # first word match (first name)
    if a.split()[0] == b.split()[0]:
        return 0.8
    return SequenceMatcher(None, a, b).ratio()

def find_best(name, hospital, candidates):
    """
    Return (doctor, score) from candidates list.
    Scoring: 70% name + 30% hospital similarity.
    """
    best, best_score = None, 0.0
    for doc in candidates:
        ns = name_score(name, doc.name)
        # hospital hint (optional boost)
        hs = SequenceMatcher(None,
             (hospital or "").lower(),
             (doc.hospital or "").lower()).ratio() if hospital and doc.hospital else 0
        score = ns * 0.75 + hs * 0.25
        if score > best_score:
            best_score, best = score, doc
    return best, best_score

def main():
    db = SessionLocal()

    # ── Find Vani B ───────────────────────────────
    vani = db.query(User).filter(User.name.ilike("%vani%")).first()
    if not vani:
        log("✗ ERROR: 'Vani B' not found in users table.")
        db.close(); return
    log(f"✓ Rep: [{vani.id}] {vani.name}\n")

    # ── Delete ALL existing investments (clean slate) ──
    deleted = db.query(Investment).delete()
    db.commit()
    log(f"  Cleared {deleted} existing investment records.\n")

    # ── Load Vani B's doctors only ────────────────
    vani_docs = db.query(Doctor).filter(
        Doctor.manager_id == vani.id,
        Doctor.is_active  != False,
    ).all()
    log(f"✓ {len(vani_docs)} doctors under Vani B\n")

    if not os.path.exists(EXCEL_PATH):
        log(f"✗ File not found: {EXCEL_PATH}")
        db.close(); return

    wb  = openpyxl.load_workbook(EXCEL_PATH)
    total_inserted = 0
    total_amount   = 0.0
    unmatched      = []

    for sheet_name, (year, month) in SHEET_MONTHS.items():
        if sheet_name not in wb.sheetnames:
            log(f"⚠ Sheet '{sheet_name}' missing — skip")
            continue

        ws = wb[sheet_name]
        log(f"══ {sheet_name}  →  {year}-{month:02d} ══")

        for row in ws.iter_rows(min_row=2, values_only=True):
            sno, doc_name, hospital, sales_val, crm_val = row

            if not sno or not doc_name:
                continue
            if str(doc_name).strip().lower() in ("total", "doctor", ""):
                continue
            if not crm_val or float(crm_val or 0) <= 0:
                continue

            amount = float(crm_val)

            doc, score = find_best(doc_name, hospital, vani_docs)

            if score >= 0.70:
                status = f"✓ [{score:.2f}]"
                matched_doc = doc
            elif score >= 0.50:
                # Low confidence — log and skip, don't create wrong links
                log(f"  ? [{score:.2f}] {doc_name} ({hospital}) → best: {doc.name if doc else '—'}  SKIPPED (low confidence)")
                unmatched.append((doc_name, hospital, year, month, amount))
                continue
            else:
                log(f"  ✗ [{score:.2f}] {doc_name} ({hospital}) → NO MATCH — will create new doctor")
                new_doc = Doctor(
                    name       = doc_name,
                    hospital   = hospital,
                    state_code = "TS",
                    zone       = "Telangana",
                    city       = "Hyderabad",
                    manager_id = vani.id,
                    is_active  = True,
                    status     = "Active",
                )
                db.add(new_doc)
                db.flush()
                vani_docs.append(new_doc)
                matched_doc = new_doc
                status = f"✚ new [{score:.2f}]"

            log(f"  {status} {doc_name} ({hospital}) → {matched_doc.name}")

            inv = Investment(
                doctor_id    = matched_doc.id,
                associate_id = vani.id,
                year         = year,
                month        = month,
                week         = 1,
                category     = InvestmentCategory.CS,
                sub_category = InvestmentSubCategory.commercial,
                amount       = amount,
                purpose      = f"CRM — {doc_name} — {hospital}",
                is_approved  = True,
            )
            db.add(inv)
            total_inserted += 1
            total_amount   += amount

        db.flush()
        log("")

    db.commit()

    log("══════════════════════════════════════")
    log(f"✓ Investments inserted : {total_inserted}")
    log(f"✓ Total amount         : ₹{total_amount:,.0f}")

    if unmatched:
        log(f"\n⚠ LOW-CONFIDENCE — NOT imported ({len(unmatched)} rows):")
        for nm, hosp, y, m, amt in unmatched:
            log(f"   {nm} | {hosp} | {y}-{m:02d} | ₹{amt:,.0f}")
        log("\n  → Check doctor names above against your DB and re-run,")
        log("    or manually add these investments from the UI.")

    db.close()

if __name__ == "__main__":
    main()
