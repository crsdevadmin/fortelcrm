"""
import_products.py — Load Fortel product price list into DB.

Usage:
    python import_products.py
    python import_products.py --xlsx path/to/other_file.xlsx

Reads: Product_Price_List_Extracted.xlsx (in same folder by default)
Columns expected: Product Name | Composition | Pack | New PTS | GST | New MRP
"""

import os, sys, re, argparse

# ── DB connection ─────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), 'backend', '.env'))

DATABASE_URL = os.getenv('DATABASE_URL', '')
m = re.match(r'postgresql://([^:]+):([^@]+)@([^/:]+)(?::(\d+))?/(.+)', DATABASE_URL)
if m:
    user, pwd, host, port, db = m.group(1), m.group(2), m.group(3), m.group(4) or '5432', m.group(5)
else:
    user, pwd, host, port, db = 'postgres', 'postgres', 'localhost', '5432', 'fortelcrm'

import psycopg2
conn = psycopg2.connect(host=host, port=int(port), database=db, user=user, password=pwd)
conn.autocommit = True
cur = conn.cursor()

# ── Read Excel ────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument('--xlsx', default=os.path.join(os.path.dirname(__file__), 'Product_Price_List_Extracted.xlsx'))
args = parser.parse_args()

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

wb = openpyxl.load_workbook(args.xlsx)
ws = wb.active

products = []
for row in ws.iter_rows(2, ws.max_row, values_only=True):
    name, composition, pack, pts, gst, mrp = row
    if not name or not pts:
        continue
    products.append({
        'name':        str(name).strip(),
        'composition': str(composition).strip() if composition else None,
        'pack':        str(pack).strip() if pack else None,
        'price':       float(pts),
        'gst':         str(gst).strip() if gst else '5%',
        'mrp':         float(mrp) if mrp else None,
    })

print(f"Found {len(products)} products in Excel\n")

inserted = updated = skipped = 0
for p in products:
    # Check if exists by name (case-insensitive)
    cur.execute("SELECT id FROM products WHERE LOWER(name) = LOWER(%s)", (p['name'],))
    row = cur.fetchone()
    if row:
        cur.execute("""
            UPDATE products SET
                composition = %s,
                pack        = %s,
                price       = %s,
                rate        = %s,
                gst         = %s,
                mrp         = %s,
                is_active   = TRUE
            WHERE id = %s
        """, (p['composition'], p['pack'], p['price'], p['price'], p['gst'], p['mrp'], row[0]))
        print(f"  UPDATED  {p['name'][:50]:<50}  PTS={p['price']:.2f}  MRP={p['mrp']}")
        updated += 1
    else:
        cur.execute("""
            INSERT INTO products (name, composition, pack, price, rate, gst, mrp, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
        """, (p['name'], p['composition'], p['pack'], p['price'], p['price'], p['gst'], p['mrp']))
        print(f"  INSERTED {p['name'][:50]:<50}  PTS={p['price']:.2f}  MRP={p['mrp']}")
        inserted += 1

conn.close()
print(f"\nDone: {inserted} inserted, {updated} updated, {skipped} skipped.")
print("Restart uvicorn to pick up model changes.")
